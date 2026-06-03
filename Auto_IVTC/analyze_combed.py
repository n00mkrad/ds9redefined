#!/usr/bin/env python3
"""
analyze_combed.py

Detects orphan-field residual combing by running two AviSynth passes and
correlating their results. Output goes to a third sheet
("Combed Frames") in frame_analysis.xlsx, and a flat DIMap.txt file.

Sheet 3 column layout (no header row; data starts at row 1):
    A : combed IVTC frame number (set by matching phase)
    B : DI-pass frame number
    C : change flag (0/1; 1 == an edit detected by MSCDetection)
    D : YDifferenceFromPrevious value, set only on edit rows + their
        previous row during edit-processing

Pass 1 (DI-Analyze.avs source + QTGMC(preset="draft"))
    Per output frame, log: frame_index, YDifferenceFromPrevious,
    change_flag where change_flag = 1 when AverageLuma(MSCDetection_clip)
    > threshold (mirrors Scenes.avs's detection criterion).

Pass 2 (IVTC-Comb.avs source)
    Per output frame, log: frame_index, IsCombedTIVTC(2).

For each row N where change == 1:
    D[N]   = YDifferenceFromPrevious of frame N+1 (i.e. ydiff at N+1)
    D[N-1] = YDifferenceFromPrevious of frame N-1 (i.e. ydiff at N-1)
    The row of {N-1, N} with the higher D wins (ties go to N).
    The losing row is deleted, plus the 5 rows before and 5 rows after
    the winner -- 10 unique rows in total per edit.

For each combed frame K from pass 2, the surviving row whose column B
is closest to K * 2.5 receives K in its column A. Matching considers
ALL surviving rows from edit-processing (winners and non-edit, non-
deletion-zone rows alike).

Finally:
    1. Rows without a column-A value are deleted.
    2. Rows without a column-D value are deleted.

Surviving rows are written to sheet "Combed Frames", and each row's
"A B" pair (space-separated, no brackets) is appended to DIMap.txt.
"""

import argparse
import bisect
import re
import subprocess
import sys
import time
from pathlib import Path

from openpyxl import Workbook, load_workbook


DI_THRESHOLD_DEFAULT = 30  # AverageLuma(sc) > N for MSCDetection-based change
SHEET_NAME = "Combed Frames"


def extract_input_line(avs_path: Path) -> str:
    pattern = re.compile(r"^\s*input\s*=", re.IGNORECASE)
    with avs_path.open("r", encoding="utf-8") as fh:
        for raw in fh:
            if pattern.match(raw):
                return raw.rstrip("\n")
    raise RuntimeError(f"Could not find 'input = ...' line in {avs_path}")


def build_di_script(input_line: str, log_path: Path, threshold: int) -> str:
    """QTGMC + MVTools MSCDetection (for change detection) plus
    YDifferenceFromPrevious (for the column-D comparison values).

    Logs (frame, ydiff, change) per frame, where:
        ydiff  = YDifferenceFromPrevious() (0.0 for frame 0)
        change = 1 when AverageLuma(MSCDetection_clip) > threshold
    """
    log_str = str(log_path).replace("\\", "/")
    return (
        f"{input_line}\n"
        f'clip = input.QTGMC(preset="draft")\n'
        f"clip\n"
        f"global super = last.MSuper(pel=2, sharp=2)\n"
        f"global vec = MAnalyse(super, isb=false, blksize=16, delta=1)\n"
        f"global sc = MSCDetection(last, vec, thSCD1=400, thSCD2=130)\n"
        f'clip2 = ScriptClip(last, """\n'
        f"    global cur_score = AverageLuma(sc)\n"
        f"    global cur_ydiff = current_frame > 0 ? YDifferenceFromPrevious() : 0.0\n"
        f'""")\n'
        f'WriteFile(clip2, "{log_str}", '
        f'"current_frame", """ "," """, '
        f'"cur_ydiff", """ "," """, '
        f'"cur_score > {threshold} ? 1 : 0", '
        f'append=false, flush=true)\n'
    )


def build_ivtc_script(input_line: str, log_path: Path) -> str:
    """Per-frame log of (frame, IsCombedTIVTC) for the IVTC-Comb source."""
    log_str = str(log_path).replace("\\", "/")
    return (
        f"{input_line}\n"
        f'WriteFile(input, "{log_str}", '
        f'"current_frame", """ "," """, '
        f'"IsCombedTIVTC(2) ? 1 : 0", '
        f'append=false, flush=true)\n'
    )


def _count_lines(path: Path) -> int:
    try:
        with path.open("rb") as fh:
            return sum(1 for _ in fh)
    except FileNotFoundError:
        return 0


def run_avs(runner_template, avs_path: Path, cwd: Path,
            log_path: Path, label: str, poll_interval: float = 0.2) -> None:
    avs_str = str(avs_path)
    if any("{AVS}" in tok for tok in runner_template):
        cmd = [tok.replace("{AVS}", avs_str) for tok in runner_template]
    else:
        cmd = runner_template + [avs_str]

    proc = subprocess.Popen(
        cmd, cwd=str(cwd),
        stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True,
    )

    use_cr = sys.stdout.isatty()
    last_count = -1
    try:
        while proc.poll() is None:
            count = _count_lines(log_path)
            if count != last_count:
                msg = f"[{label}] frame {count}"
                if use_cr:
                    sys.stdout.write("\r" + msg + " " * 8)
                else:
                    sys.stdout.write(msg + "\n")
                sys.stdout.flush()
                last_count = count
            time.sleep(poll_interval)
    finally:
        _, stderr = proc.communicate()

    final = _count_lines(log_path)
    end_msg = f"[{label}] frame {final} (done)"
    if use_cr:
        sys.stdout.write("\r" + end_msg + " " * 8 + "\n")
    else:
        sys.stdout.write(end_msg + "\n")
    sys.stdout.flush()

    if proc.returncode != 0:
        raise RuntimeError(
            f"AVS runner failed for {avs_path.name}\n"
            f"cmd: {' '.join(cmd)}\n"
            f"stderr:\n{stderr}"
        )


def parse_di_log(log_path: Path) -> list[dict]:
    """Parse 'frame,ydiff,change' lines into a list of row dicts."""
    rows: list[dict] = []
    with log_path.open("r", encoding="utf-8") as fh:
        for line in fh:
            parts = [p.strip() for p in line.strip().split(",")]
            if len(parts) != 3:
                continue
            try:
                frame = int(parts[0])
                ydiff = float(parts[1])
                change = int(parts[2])
            except ValueError:
                continue
            rows.append({
                "frame": frame, "ydiff": ydiff, "change": change,
                "d": None, "a": None,
            })
    rows.sort(key=lambda r: r["frame"])
    return rows


def parse_ivtc_log(log_path: Path) -> list[int]:
    combed: list[int] = []
    with log_path.open("r", encoding="utf-8") as fh:
        for line in fh:
            parts = [p.strip() for p in line.strip().split(",")]
            if len(parts) != 2:
                continue
            try:
                frame = int(parts[0])
                flag = int(parts[1])
            except ValueError:
                continue
            if flag == 1:
                combed.append(frame)
    return combed


def process_edits(rows: list[dict]) -> list[dict]:
    """For each row N where change==1:
        D[N]   = ydiff at frame N+1
        D[N-1] = ydiff at frame N-1
    Compare D[N] and D[N-1]; the higher wins (ties go to row N).
    Mark for deletion: loser + 5 rows before and 5 rows after the winner
    (10 unique rows when the loser overlaps with the surrounding zone).

    All deletions are unioned across edits, then applied in one pass.
    Returns the surviving rows.
    """
    n = len(rows)
    deleted: set[int] = set()
    for i in range(n):
        if rows[i]["change"] != 1:
            continue
        if i - 1 < 0 or i + 1 >= n:
            continue  # need both neighbors for D values

        d_at_N = rows[i + 1]["ydiff"]      # YDiffPrev at frame N+1
        d_at_Nm1 = rows[i - 1]["ydiff"]    # YDiffPrev at frame N-1
        rows[i]["d"] = d_at_N
        rows[i - 1]["d"] = d_at_Nm1

        if d_at_N >= d_at_Nm1:
            winner = i
            loser = i - 1
        else:
            winner = i - 1
            loser = i

        deleted.add(loser)
        for offset in (-5, -4, -3, -2, -1, 1, 2, 3, 4, 5):
            k = winner + offset
            if 0 <= k < n:
                deleted.add(k)

    return [r for idx, r in enumerate(rows) if idx not in deleted]


def assign_combed_to_rows(rows: list[dict], combed_frames: list[int]) -> None:
    """For each combed IVTC frame K, find the row whose column B is
    closest to K * 2.5 and write K into that row's column A.

    Considers ALL provided rows as match candidates (per the literal
    spec: matching is over all surviving rows from edit-processing,
    not only winners). Non-winner matches will be culled later by the
    final 'delete without D' filter."""
    if not rows:
        return
    rows_sorted = sorted(rows, key=lambda r: r["frame"])
    b_values = [r["frame"] for r in rows_sorted]
    for k in combed_frames:
        target = k * 2.5
        idx = bisect.bisect_left(b_values, target)
        cands = []
        if idx > 0:
            cands.append(idx - 1)
        if idx < len(b_values):
            cands.append(idx)
        if not cands:
            continue
        best = min(cands, key=lambda i: abs(b_values[i] - target))
        rows_sorted[best]["a"] = k


def default_runner() -> list[str]:
    return [
        "ffmpeg", "-hide_banner", "-loglevel", "error", "-y",
        "-i", "{AVS}", "-f", "null", "-",
    ]


def parse_runner(value: str) -> list[str]:
    return value.split()


def write_sheet(xlsx_path: Path, rows: list[dict]) -> None:
    """Write final filtered rows to Sheet 3 using columns A-D only.
        A = combed IVTC frame, B = frame, C = change flag, D = ydiff
    No header row; data starts at row 1.
    """
    if xlsx_path.exists():
        wb = load_workbook(xlsx_path)
    else:
        wb = Workbook()
        default = wb.active
        wb.remove(default)

    if SHEET_NAME in wb.sheetnames:
        del wb[SHEET_NAME]
    ws = wb.create_sheet(SHEET_NAME)

    for i, r in enumerate(rows, start=1):
        ws.cell(row=i, column=1, value=r["a"])
        ws.cell(row=i, column=2, value=r["frame"])
        ws.cell(row=i, column=3, value=r["change"])
        ws.cell(row=i, column=4, value=r["d"])

    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(xlsx_path)


def write_dimap(dimap_path: Path, rows: list[dict]) -> None:
    with dimap_path.open("w", encoding="utf-8") as fh:
        for r in rows:
            fh.write(f"{r['a']} {r['frame']}\n")


def main() -> int:
    parser = argparse.ArgumentParser(
        description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument("--project-dir", type=Path,
                        default=Path(__file__).resolve().parent,
                        help="Folder with DI-Analyze.avs and IVTC-Comb.avs "
                             "(default: script dir)")
    parser.add_argument("--xlsx", type=Path, default=None,
                        help="Workbook to write into "
                             "(default: <project>/frame_analysis.xlsx)")
    parser.add_argument("--runner", type=parse_runner, default=None,
                        help='AVS runner command with {AVS} placeholder, e.g. '
                             '"ffmpeg -hide_banner -loglevel error -y -i {AVS} '
                             '-f null -"')
    parser.add_argument("--threshold", type=int, default=DI_THRESHOLD_DEFAULT,
                        help=f"AverageLuma(sc) change threshold for DI pass "
                             f"(default: {DI_THRESHOLD_DEFAULT})")
    parser.add_argument("--keep-temp", action="store_true",
                        help="Keep the generated AVS and log files for inspection")
    args = parser.parse_args()

    project_dir: Path = args.project_dir.resolve()
    di_avs = project_dir / "DI-Analyze.avs"
    ivtc_avs = project_dir / "IVTC-Comb.avs"
    for f in (di_avs, ivtc_avs):
        if not f.exists():
            print(f"Missing required AVS: {f}", file=sys.stderr)
            return 1

    xlsx_path: Path = (args.xlsx or project_dir / "frame_analysis.xlsx").resolve()
    runner_cmd = args.runner or default_runner()

    di_input = extract_input_line(di_avs)
    ivtc_input = extract_input_line(ivtc_avs)

    di_avs_file = project_dir / "_di_analyze.avs"
    di_log = project_dir / "_di_analyze.log"
    ivtc_avs_file = project_dir / "_ivtc_comb.avs"
    ivtc_log = project_dir / "_ivtc_comb.log"
    for lf in (di_log, ivtc_log):
        if lf.exists():
            lf.unlink()

    di_avs_file.write_text(
        build_di_script(di_input, di_log, args.threshold), encoding="utf-8")
    ivtc_avs_file.write_text(
        build_ivtc_script(ivtc_input, ivtc_log), encoding="utf-8")

    temp_files = [di_avs_file, di_log, ivtc_avs_file, ivtc_log]

    try:
        # Pass 1
        run_avs(runner_cmd, di_avs_file, project_dir, di_log, "DI")
        # Pass 2
        run_avs(runner_cmd, ivtc_avs_file, project_dir, ivtc_log, "IVTC")

        rows = parse_di_log(di_log)
        if not rows:
            print("No DI frames analyzed; aborting.", file=sys.stderr)
            return 2

        n_changes = sum(1 for r in rows if r["change"] == 1)
        survivors = process_edits(rows)
        n_with_d = sum(1 for r in survivors if r["d"] is not None)
        print(f"DI rows: {len(rows)}  change frames: {n_changes}  "
              f"survivors after edit-processing: {len(survivors)} "
              f"(of which {n_with_d} have a D value)")

        combed = parse_ivtc_log(ivtc_log)
        print(f"IVTC combed frames: {len(combed)}")

        assign_combed_to_rows(survivors, combed)

        # Final filters: drop rows without A, then drop rows without D.
        with_a = [r for r in survivors if r["a"] is not None]
        final = [r for r in with_a if r["d"] is not None]
        final.sort(key=lambda r: r["a"])
        print(f"Combed-assigned rows: {len(with_a)}  "
              f"surviving 'no D' filter: {len(final)}")

        write_sheet(xlsx_path, final)
        dimap_path = project_dir / "DIMap.txt"
        write_dimap(dimap_path, final)

        print(f"Saved: {xlsx_path}  (sheet: {SHEET_NAME!r})")
        print(f"Saved: {dimap_path}")
    finally:
        if not args.keep_temp:
            for f in temp_files:
                try:
                    f.unlink()
                except FileNotFoundError:
                    pass

    return 0


if __name__ == "__main__":
    sys.exit(main())
