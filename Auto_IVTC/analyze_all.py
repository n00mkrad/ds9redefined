#!/usr/bin/env python3
"""
analyze_all.py

Self-contained driver that produces frame_analysis.xlsx in one pass:

    Sheet 1 "Frame Analysis"  -- PD1..PD5 combed/clean status per frame
    Sheet 2 "Scene Ranges"    -- start/end frame pairs for each scene
                                 (scene changes + commercial-break starts
                                 from Breaks.txt are merged as splits)

AVS scripting:
  * Pulldown.avs:    used for source line; combing checked with
                     IsCombedTIVTC(4) (threshold matches Pulldown.avs).
  * Scenes.avs:      used for source line; scene detection uses MVTools
                     MSCDetection on QTGMC + TDecimate output, matching
                     the current Scenes.avs (MSuper/MAnalyse/MSCDetection
                     then AverageLuma > 30).

Both AVS files have their `input = ...` line read at runtime, so source
changes in either file are picked up automatically.
"""

import argparse
import bisect  # noqa: F401  (kept available for downstream tools)
import re
import subprocess
import sys
import time
from pathlib import Path

from openpyxl import Workbook


# --- Pulldown phase constants ----------------------------------------------

PD_PARAMS = {
    "PD1": "0, 2",
    "PD2": "0, 3",
    "PD3": "1, 3",
    "PD4": "1, 4",
    "PD5": "2, 4",
}
COMBED_CTHRESH = 4  # IsCombedTIVTC threshold (Pulldown.avs uses 4)

# --- Scene-detection constants (mirror Scenes.avs) -------------------------

SCENE_SCORE_THRESHOLD = 30   # AverageLuma(sc) > 30 == scene change
MS_PEL = 2
MS_SHARP = 2
MA_BLKSIZE = 16
MA_DELTA = 1
MSCD_THSCD1 = 400
MSCD_THSCD2 = 130

# --- Sheet names -----------------------------------------------------------

SHEET1_NAME = "Frame Analysis"
SHEET2_NAME = "Scene Ranges"


# --------------------------------------------------------------------------
# Shared utilities
# --------------------------------------------------------------------------

def extract_input_line(avs_path: Path) -> str:
    """Return the first `input = ...` assignment line from the given AVS."""
    pattern = re.compile(r"^\s*input\s*=", re.IGNORECASE)
    with avs_path.open("r", encoding="utf-8") as fh:
        for raw in fh:
            if pattern.match(raw):
                return raw.rstrip("\n")
    raise RuntimeError(f"Could not find 'input = ...' line in {avs_path}")


def _count_lines(path: Path) -> int:
    try:
        with path.open("rb") as fh:
            return sum(1 for _ in fh)
    except FileNotFoundError:
        return 0


def run_avs(runner_template: list[str], avs_path: Path, cwd: Path,
            log_path: Path, label: str, poll_interval: float = 0.2) -> None:
    """Run the AVS through the configured runner; show a live frame counter."""
    avs_str = str(avs_path)
    if any("{AVS}" in tok for tok in runner_template):
        cmd = [tok.replace("{AVS}", avs_str) for tok in runner_template]
    else:
        cmd = runner_template + [avs_str]

    proc = subprocess.Popen(
        cmd, cwd=str(cwd),
        stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True,
    )

    use_cr = sys.stdout.isatty()
    last_count = -1
    try:
        while proc.poll() is None:
            count = _count_lines(log_path)
            if count != last_count:
                msg = f"[{label}] frame {count}"
                if use_cr:
                    sys.stdout.write("\r" + msg + " " * 8)
                else:
                    sys.stdout.write(msg + "\n")
                sys.stdout.flush()
                last_count = count
            time.sleep(poll_interval)
    finally:
        _, stderr = proc.communicate()

    final = _count_lines(log_path)
    end_msg = f"[{label}] frame {final} (done)"
    if use_cr:
        sys.stdout.write("\r" + end_msg + " " * 8 + "\n")
    else:
        sys.stdout.write(end_msg + "\n")
    sys.stdout.flush()

    if proc.returncode != 0:
        raise RuntimeError(
            f"AVS runner failed for {avs_path.name}\n"
            f"cmd: {' '.join(cmd)}\n"
            f"stderr:\n{stderr}"
        )


def default_runner() -> list[str]:
    return [
        "ffmpeg", "-hide_banner", "-loglevel", "error", "-y",
        "-i", "{AVS}", "-f", "null", "-",
    ]


def parse_runner(value: str) -> list[str]:
    return value.split()


# --------------------------------------------------------------------------
# Pulldown phase
# --------------------------------------------------------------------------

def build_pd_script(input_line: str, pd_args: str, log_path: Path) -> str:
    """One AVS per PD variant; writes 'frame,flag' per output frame."""
    log_str = str(log_path).replace("\\", "/")
    return (
        f"{input_line}\n"
        f"inputDW = input.DoubleWeave()\n"
        f"clip = inputDW.Pulldown({pd_args})\n"
        f'WriteFile(clip, "{log_str}", "current_frame", """ "," """, '
        f'"IsCombedTIVTC({COMBED_CTHRESH}) ? 0 : 1", '
        f'append=false, flush=true)\n'
    )


def parse_pd_log(log_path: Path) -> dict[int, int]:
    """Parse 'frame,flag' lines into {frame_index: 0|1}."""
    results: dict[int, int] = {}
    with log_path.open("r", encoding="utf-8") as fh:
        for line in fh:
            line = line.strip()
            if not line:
                continue
            parts = [p.strip() for p in line.split(",")]
            if len(parts) != 2:
                continue
            try:
                frame = int(parts[0])
                flag = int(parts[1])
            except ValueError:
                continue
            results[frame] = flag
    return results


# --------------------------------------------------------------------------
# Scenes phase
# --------------------------------------------------------------------------

def build_scenes_script(input_line: str, log_path: Path,
                        score_threshold: int) -> str:
    """Generate AVS mirroring Scenes.avs (MSCDetection-based) but with
    WriteFile producing 'frame,is_scene_change' lines."""
    log_str = str(log_path).replace("\\", "/")
    return (
        f"{input_line}\n"
        f'DI = input.QTGMC(preset="draft").TDecimate(mode=2, rate=23.976)\n'
        f"DI\n"
        f"global super = last.MSuper(pel={MS_PEL}, sharp={MS_SHARP})\n"
        f"global vec = MAnalyse(super, isb=false, "
        f"blksize={MA_BLKSIZE}, delta={MA_DELTA})\n"
        f"global sc = MSCDetection(last, vec, "
        f"thSCD1={MSCD_THSCD1}, thSCD2={MSCD_THSCD2})\n"
        f'clip = ScriptClip(last, """\n'
        f"    global cur_score = AverageLuma(sc)\n"
        f'""")\n'
        f'WriteFile(clip, "{log_str}", "current_frame", """ "," """, '
        f'"cur_score > {score_threshold} ? 1 : 0", '
        f'append=false, flush=true)\n'
    )


def parse_scenes_log(log_path: Path) -> tuple[list[int], int]:
    """Return (scene_change_frames_0based, last_frame_0based)."""
    scenes: list[int] = []
    last_frame = -1
    with log_path.open("r", encoding="utf-8") as fh:
        for line in fh:
            line = line.strip()
            if not line:
                continue
            parts = [p.strip() for p in line.split(",")]
            if len(parts) != 2:
                continue
            try:
                frame = int(parts[0])
                flag = int(parts[1])
            except ValueError:
                continue
            if frame > last_frame:
                last_frame = frame
            if flag == 1:
                scenes.append(frame)
    scenes.sort()
    return scenes, last_frame


def load_breaks(breaks_path: Path) -> list[int]:
    """Frame numbers (one per line) from Breaks.txt; blanks and #-comments
    ignored. Non-integer lines are skipped with a warning."""
    if not breaks_path.exists():
        return []
    frames: list[int] = []
    with breaks_path.open("r", encoding="utf-8") as fh:
        for lineno, raw in enumerate(fh, start=1):
            line = raw.strip()
            if not line or line.startswith("#"):
                continue
            try:
                frames.append(int(line))
            except ValueError:
                print(f"WARNING: Breaks.txt line {lineno}: ignoring "
                      f"non-integer value {line!r}", file=sys.stderr)
    return frames


# --------------------------------------------------------------------------
# Workbook builders
# --------------------------------------------------------------------------

def populate_frame_analysis(ws, pd_results: dict[str, dict[int, int]]) -> None:
    """Sheet 1: Frame, PD1..PD5. Row 1 is the header."""
    ws.cell(row=1, column=1, value="Frame")
    for col_idx, pd_name in enumerate(PD_PARAMS, start=2):
        ws.cell(row=1, column=col_idx, value=pd_name)

    if not any(pd_results.values()):
        return

    max_frame = max(max(r.keys()) for r in pd_results.values() if r)
    for frame in range(max_frame + 1):
        row = frame + 2
        ws.cell(row=row, column=1, value=frame)
        for col_idx, pd_name in enumerate(PD_PARAMS, start=2):
            ws.cell(row=row, column=col_idx,
                    value=pd_results[pd_name].get(frame))


def populate_scene_ranges(ws, scenes: list[int], breaks: list[int],
                          last_frame: int) -> None:
    """Sheet 2: 0-based ranges, A1 = 0. Breaks merge in as additional
    range starts identical in role to scene changes."""
    valid_breaks = [b for b in breaks if 0 < b <= last_frame]
    dropped = len(breaks) - len(valid_breaks)
    if dropped:
        print(f"WARNING: {dropped} break frame(s) outside valid range "
              f"(1..{last_frame}) were ignored", file=sys.stderr)

    all_splits = sorted(set(scenes) | set(valid_breaks))

    row = 1
    ws.cell(row=row, column=1, value=0)
    for sc in all_splits:
        ws.cell(row=row, column=2, value=sc - 1)
        row += 1
        ws.cell(row=row, column=1, value=sc)
    ws.cell(row=row, column=2, value=last_frame)


# --------------------------------------------------------------------------
# Main
# --------------------------------------------------------------------------

def main() -> int:
    parser = argparse.ArgumentParser(
        description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument("--project-dir", type=Path,
                        default=Path(__file__).resolve().parent,
                        help="Folder with Pulldown.avs / Scenes.avs / "
                             "Breaks.txt (default: script dir)")
    parser.add_argument("--xlsx", type=Path, default=None,
                        help="Output workbook "
                             "(default: <project>/frame_analysis.xlsx)")
    parser.add_argument("--runner", type=parse_runner, default=None,
                        help='AVS runner with {AVS} placeholder, e.g. '
                             '"ffmpeg -hide_banner -loglevel error -y -i {AVS} '
                             '-f null -"')
    parser.add_argument("--scene-threshold", type=int,
                        default=SCENE_SCORE_THRESHOLD,
                        help=f"AverageLuma(sc) > N triggers a scene change "
                             f"(default: {SCENE_SCORE_THRESHOLD})")
    parser.add_argument("--breaks", type=Path, default=None,
                        help="Path to Breaks.txt "
                             "(default: <project>/Breaks.txt)")
    parser.add_argument("--keep-temp", action="store_true",
                        help="Keep generated AVS and log files for inspection")
    args = parser.parse_args()

    project_dir: Path = args.project_dir.resolve()
    pulldown_avs = project_dir / "Pulldown.avs"
    scenes_avs = project_dir / "Scenes.avs"
    for f in (pulldown_avs, scenes_avs):
        if not f.exists():
            print(f"Missing required AVS: {f}", file=sys.stderr)
            return 1

    xlsx_path: Path = (args.xlsx or project_dir / "frame_analysis.xlsx").resolve()
    runner_cmd = args.runner or default_runner()

    pd_input = extract_input_line(pulldown_avs)
    sc_input = extract_input_line(scenes_avs)

    # ---- Phase 1: Pulldown analysis (5 PD passes) ----
    pd_results: dict[str, dict[int, int]] = {}
    temp_files: list[Path] = []
    try:
        for pd_name, pd_args in PD_PARAMS.items():
            avs_file = project_dir / f"_analysis_{pd_name}.avs"
            log_file = project_dir / f"_analysis_{pd_name}.log"
            if log_file.exists():
                log_file.unlink()
            avs_file.write_text(
                build_pd_script(pd_input, pd_args, log_file),
                encoding="utf-8",
            )
            temp_files.extend([avs_file, log_file])

            run_avs(runner_cmd, avs_file, cwd=project_dir,
                    log_path=log_file, label=pd_name)
            if not log_file.exists():
                raise RuntimeError(
                    f"No log produced for {pd_name}; AVS likely did not "
                    f"process any frames."
                )
            pd_results[pd_name] = parse_pd_log(log_file)

        # ---- Phase 2: Scene analysis (single pass) ----
        sc_avs_file = project_dir / "_scene_analysis.avs"
        sc_log_file = project_dir / "_scene_analysis.log"
        if sc_log_file.exists():
            sc_log_file.unlink()
        sc_avs_file.write_text(
            build_scenes_script(sc_input, sc_log_file, args.scene_threshold),
            encoding="utf-8",
        )
        temp_files.extend([sc_avs_file, sc_log_file])

        run_avs(runner_cmd, sc_avs_file, cwd=project_dir,
                log_path=sc_log_file, label="scenes")
        if not sc_log_file.exists():
            raise RuntimeError("No log produced for scene analysis.")

        scenes, last_frame = parse_scenes_log(sc_log_file)
        if last_frame < 0:
            print("No scene-analysis frames returned; aborting.",
                  file=sys.stderr)
            return 2

        breaks_path = args.breaks or (project_dir / "Breaks.txt")
        breaks = load_breaks(breaks_path)

        # ---- Phase 3: Build workbook ----
        wb = Workbook()
        ws1 = wb.active
        ws1.title = SHEET1_NAME
        populate_frame_analysis(ws1, pd_results)

        ws2 = wb.create_sheet(SHEET2_NAME)
        populate_scene_ranges(ws2, scenes, breaks, last_frame)

        xlsx_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(xlsx_path)

        # ---- Summary ----
        print()
        for pd_name in PD_PARAMS:
            print(f"  {pd_name}: {len(pd_results[pd_name])} frames")
        print(f"  scene changes: {len(scenes)} "
              f"(score threshold > {args.scene_threshold})")
        if breaks_path.exists():
            print(f"  breaks loaded: {len(breaks)} from {breaks_path.name}")
        else:
            print(f"  breaks file not found at {breaks_path} (skipped)")
        print(f"  last frame (0-based): {last_frame}")
        print(f"  saved: {xlsx_path}")
    finally:
        if not args.keep_temp:
            for f in temp_files:
                try:
                    f.unlink()
                except FileNotFoundError:
                    pass

    return 0


if __name__ == "__main__":
    sys.exit(main())
