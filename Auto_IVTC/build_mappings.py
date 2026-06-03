#!/usr/bin/env python3
"""
build_mappings.py

Reads frame_analysis.xlsx (produced by analyze_pulldowns.py and
analyze_scenes.py) and fills columns C-H of the "Scene Ranges" sheet,
then writes per-pulldown mapping text files.

For every existing row in "Scene Ranges" with start frame A in column A
and end frame B in column B:

    Column C = SUM of Sheet 1 column B (PD1) over frames A..B  (clean count)
    Column D = SUM of Sheet 1 column C (PD2) over frames A..B
    Column E = SUM of Sheet 1 column D (PD3) over frames A..B
    Column F = SUM of Sheet 1 column E (PD4) over frames A..B
    Column G = SUM of Sheet 1 column F (PD5) over frames A..B
    Column H = name of the winning pulldown (PD1..PD5)

Tiebreaker rules for column H when two or more of C..G tie at the max:
    C == D -> PD1
    D == E -> PD2
    E == F -> PD3
    F == G -> PD4
    G == C -> PD5

Rules are applied in the order listed above; the first one whose pair is
present in the set of tied-max columns wins. Non-adjacent or 3+ way ties
that no rule covers fall back to the leftmost tied column.

Finally, each row's range is written to its assigned mapping text file
(PD1Map.txt..PD5Map.txt) on its own line in the format:

    [A B]
"""

import argparse
import sys
from pathlib import Path

from openpyxl import load_workbook


SHEET1_NAME = "Frame Analysis"
SHEET2_NAME = "Scene Ranges"

PD_NAMES = ("PD1", "PD2", "PD3", "PD4", "PD5")
COL_TO_PD = {"C": "PD1", "D": "PD2", "E": "PD3", "F": "PD4", "G": "PD5"}


def col_idx(letter: str) -> int:
    return ord(letter.upper()) - ord("A") + 1


def build_frame_lookup(ws1) -> dict[int, tuple[int, int, int, int, int]]:
    """Return {frame_number: (PD1, PD2, PD3, PD4, PD5)} from Sheet 1.

    Sheet 1 layout: row 1 = headers, row 2.. = data
        col A = Frame, col B..F = PD1..PD5 (0=combed, 1=clean)
    Missing cells are treated as 0.
    """
    lookup: dict[int, tuple[int, int, int, int, int]] = {}
    for row in ws1.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        try:
            frame = int(row[0])
        except (TypeError, ValueError):
            continue
        vals = []
        for i in range(1, 6):
            v = row[i] if i < len(row) else None
            try:
                vals.append(int(v) if v is not None else 0)
            except (TypeError, ValueError):
                vals.append(0)
        lookup[frame] = tuple(vals)  # type: ignore[assignment]
    return lookup


def pick_winner(c: int, d: int, e: int, f: int, g: int) -> str:
    """Choose PD1..PD5 by max value with the documented tiebreaker rules."""
    vals = {"C": c, "D": d, "E": e, "F": f, "G": g}
    mx = max(vals.values())
    tied = {k for k, v in vals.items() if v == mx}

    if len(tied) == 1:
        return COL_TO_PD[next(iter(tied))]

    rules = (
        ({"C", "D"}, "PD1"),
        ({"D", "E"}, "PD2"),
        ({"E", "F"}, "PD3"),
        ({"F", "G"}, "PD4"),
        ({"G", "C"}, "PD5"),
    )
    for pair, pd_name in rules:
        if pair.issubset(tied):
            return pd_name

    # No adjacent-pair rule matched (e.g. {C, E}); fall back to leftmost.
    leftmost = min(tied)  # lexicographic min among C..G == leftmost column
    return COL_TO_PD[leftmost]


def main() -> int:
    parser = argparse.ArgumentParser(
        description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument("--project-dir", type=Path,
                        default=Path(__file__).resolve().parent,
                        help="Folder for output text files (default: script dir)")
    parser.add_argument("--xlsx", type=Path, default=None,
                        help="Workbook to read/update "
                             "(default: <project>/frame_analysis.xlsx)")
    args = parser.parse_args()

    project_dir: Path = args.project_dir.resolve()
    xlsx_path: Path = (args.xlsx or project_dir / "frame_analysis.xlsx").resolve()

    if not xlsx_path.exists():
        print(f"Workbook not found: {xlsx_path}", file=sys.stderr)
        return 1

    wb = load_workbook(xlsx_path)
    for needed in (SHEET1_NAME, SHEET2_NAME):
        if needed not in wb.sheetnames:
            print(f"Missing sheet {needed!r} in {xlsx_path}", file=sys.stderr)
            return 2

    ws1 = wb[SHEET1_NAME]
    ws2 = wb[SHEET2_NAME]

    frames = build_frame_lookup(ws1)
    if not frames:
        print(f"No frame data found in sheet {SHEET1_NAME!r}", file=sys.stderr)
        return 3

    project_dir.mkdir(parents=True, exist_ok=True)
    mapping_paths = {pd: project_dir / f"{pd}Map.txt" for pd in PD_NAMES}
    mapping_files = {pd: p.open("w", encoding="utf-8") for pd, p in mapping_paths.items()}
    row_counts = {pd: 0 for pd in PD_NAMES}

    try:
        for row_idx in range(1, ws2.max_row + 1):
            a_val = ws2.cell(row=row_idx, column=1).value
            b_val = ws2.cell(row=row_idx, column=2).value
            if a_val is None or b_val is None:
                continue
            try:
                a = int(a_val)
                b = int(b_val)
            except (TypeError, ValueError):
                continue
            if b < a:
                continue

            # Sum each PD column across the range A..B (inclusive)
            sums = [0, 0, 0, 0, 0]
            for f in range(a, b + 1):
                entry = frames.get(f)
                if entry is None:
                    continue
                for i in range(5):
                    sums[i] += entry[i]

            pd1, pd2, pd3, pd4, pd5 = sums
            ws2.cell(row=row_idx, column=col_idx("C"), value=pd1)
            ws2.cell(row=row_idx, column=col_idx("D"), value=pd2)
            ws2.cell(row=row_idx, column=col_idx("E"), value=pd3)
            ws2.cell(row=row_idx, column=col_idx("F"), value=pd4)
            ws2.cell(row=row_idx, column=col_idx("G"), value=pd5)

            winner = pick_winner(pd1, pd2, pd3, pd4, pd5)
            ws2.cell(row=row_idx, column=col_idx("H"), value=winner)

            mapping_files[winner].write(f"[{a} {b}]\n")
            row_counts[winner] += 1
    finally:
        for fh in mapping_files.values():
            fh.close()

    wb.save(xlsx_path)

    print(f"Updated {xlsx_path}  (sheet: {SHEET2_NAME!r})")
    for pd in PD_NAMES:
        print(f"  {mapping_paths[pd].name}: {row_counts[pd]} range(s)")

    return 0


if __name__ == "__main__":
    sys.exit(main())
